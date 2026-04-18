import openpyxl
import json

excel_path = r'C:\Users\shudg\Downloads\1殿生編號+職務20260418.xlsx'
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['Page1']

data = []
# Assuming Dharma Names start from row 1, column B (2)
# Titles start from column C (3) onwards
for row in range(1, sheet.max_row + 1):
    dharma_name = sheet.cell(row=row, column=2).value
    if not dharma_name:
        continue
    
    # Filter out numeric only dharma names if they are just IDs (optional check)
    if isinstance(dharma_name, (int, float)):
        # If it's the sequence number column handled wrongly, we'll see
        # But looking at screenshot, B is the Name.
        pass

    groups = []
    # Loop through columns C (3) to the end
    for col in range(3, sheet.max_column + 1):
        val = sheet.cell(row=row, column=col).value
        # Ignore "職務" headers or empty values
        if val and val != "職務" and not str(val).isdigit():
            groups.append(str(val).strip())
    
    if groups:
        data.append({
            "name": str(dharma_name).strip(),
            "groups": groups
        })

# Save as JSON for next step
with open('scratch/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print(f"Successfully parsed {len(data)} entries.")
